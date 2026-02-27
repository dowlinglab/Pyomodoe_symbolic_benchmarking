#!/usr/bin/env python
"""
Summarize DoE benchmark JSONs into an Excel workbook.

Usage:
  python summarize_doe_jsons_excel.py --root <dir> --out <file.xlsx>

Behavior:
- Recursively scans root for JSON files.
- Groups JSONs by their parent directory (case folder).
- Selects one CENTRAL and one SYMBOLIC JSON per case based on filename keywords.
- Writes one sheet per case with metrics, differences, and ratios.
"""

import argparse
import json
import os
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font


METRIC_KEYS = [
    "doe_build_time",
    "doe_init_time",
    "doe_solve_time",
    "ipopt_cpu_nlp_feval",
    "ipopt_cpu_wo_feval",
    "doe_wall_time",
    "ipopt_iters",
    "obj_eval",
    "grad_eval",
    "eq_con_eval",
    "eq_jac_eval",
    "hess_eval",
]

METRIC_LABELS = [
    "Build",
    "Init.",
    "Solve",
    "NLP eval.",
    "CPU",
    "Wall clck.",
    "Iterations",
    "Obj. fun.",
    "Obj. grad.",
    "Eq. constr.",
    "Eq. jac.",
    "Lag. Hess.",
]

# Metrics 1..6 are times, 7..12 are counts.
TIME_KEYS = set(METRIC_KEYS[:6])
COUNT_KEYS = set(METRIC_KEYS[6:])


def parse_args() -> argparse.Namespace:
    """Parse CLI arguments."""
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", required=True, help="Root directory to scan for JSON files")
    parser.add_argument("--out", required=True, help="Output XLSX file path")
    return parser.parse_args()


def find_jsons(root: Path) -> Dict[Path, List[Path]]:
    """
    Recursively find JSON files and group by case directory.

    Case directory is defined as:
    - If JSON is under <case>/central or <case>/symbolic, group by <case>
    - Otherwise, group by the JSON's parent directory
    """
    groups: Dict[Path, List[Path]] = {}
    for path in root.rglob("*.json"):
        if not path.is_file():
            continue
        parent = path.parent
        if parent.name.lower() in {"central", "symbolic"} and parent.parent != parent:
            case_dir = parent.parent
        else:
            case_dir = parent
        groups.setdefault(case_dir, []).append(path)
    return groups


def pick_candidate(paths: List[Path], keywords: List[str]) -> List[Path]:
    """Filter list of paths by case-insensitive keyword match in filename."""
    out = []
    for p in paths:
        name = p.name.lower()
        if any(k in name for k in keywords):
            out.append(p)
    return out


def choose_shortest(paths: List[Path]) -> Path:
    """Deterministically choose the shortest filename, tiebreak by name."""
    return sorted(paths, key=lambda p: (len(p.name), p.name))[0]


def load_json(path: Path) -> Dict:
    """Load JSON file into dict."""
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def to_float(value):
    """Best-effort numeric conversion; return None if not numeric."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value)
        except ValueError:
            return None
    return None


def set_col_widths(ws, widths: Dict[int, int]) -> None:
    """Set worksheet column widths (1-based index)."""
    for col_idx, width in widths.items():
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = width


def main() -> None:
    args = parse_args()
    root = Path(args.root)
    out_path = Path(args.out)

    if not root.exists():
        raise SystemExit(f"Root path not found: {root}")

    case_groups = find_jsons(root)
    case_dirs = sorted(case_groups.keys(), key=lambda p: p.name)

    wb = Workbook()
    # Remove default sheet to control ordering; add a placeholder if needed later.
    default_ws = wb.active
    wb.remove(default_ws)

    case_index = 1
    for case_dir in case_dirs:
        json_paths = case_groups[case_dir]

        central_matches = pick_candidate(json_paths, ["central"])
        symbolic_matches = pick_candidate(json_paths, ["sym", "symbolic", "pynumero"])

        if not central_matches or not symbolic_matches:
            print(f"WARNING: skipping {case_dir} (missing central or symbolic JSON)")
            continue

        central_path = choose_shortest(central_matches)
        symbolic_path = choose_shortest(symbolic_matches)

        if len(central_matches) > 1:
            print(f"INFO: {case_dir} central candidates: {[p.name for p in central_matches]}")
            print(f"INFO: selected central: {central_path.name}")
        if len(symbolic_matches) > 1:
            print(f"INFO: {case_dir} symbolic candidates: {[p.name for p in symbolic_matches]}")
            print(f"INFO: selected symbolic: {symbolic_path.name}")

        central = load_json(central_path)
        symbolic = load_json(symbolic_path)

        # Sheet name must be <= 31 chars.
        sheet_name = f"case-{case_index}"
        sheet_name = sheet_name[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Header info above table.
        ws.cell(row=1, column=1, value=f"Case folder: {case_dir.name}")
        ws.cell(row=2, column=1, value=f"Central JSON: {central_path.name}")
        ws.cell(row=3, column=1, value=f"Symbolic JSON: {symbolic_path.name}")

        # Table header (row 5).
        header_row = 5
        headers = ["Row", "Metric", "central", "symbolic", "diff", "ratio"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True)

        # Freeze panes at header row (keep header visible).
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        # Data rows 6..17
        for i, (key, label) in enumerate(zip(METRIC_KEYS, METRIC_LABELS), start=1):
            row = header_row + i
            c_val = central.get(key, "NA")
            s_val = symbolic.get(key, "NA")

            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=label)
            ws.cell(row=row, column=3, value=c_val)
            ws.cell(row=row, column=4, value=s_val)

            # Compute diff/ratio if numeric.
            c_num = to_float(c_val)
            s_num = to_float(s_val)
            diff_val = (s_num - c_num) if (s_num is not None and c_num is not None) else ""
            ratio_val = (s_num / c_num) if (s_num is not None and c_num not in (None, 0.0)) else ""

            ws.cell(row=row, column=5, value=diff_val)
            ws.cell(row=row, column=6, value=ratio_val)

            # Apply number formats by metric type.
            if key in TIME_KEYS:
                for col in (3, 4, 5):
                    if isinstance(ws.cell(row=row, column=col).value, (int, float)):
                        ws.cell(row=row, column=col).number_format = "0.00"
            if key in COUNT_KEYS:
                for col in (3, 4, 5):
                    if isinstance(ws.cell(row=row, column=col).value, (int, float)):
                        ws.cell(row=row, column=col).number_format = "0"
            # Ratio always in column 6 if numeric.
            if isinstance(ws.cell(row=row, column=6).value, (int, float)):
                ws.cell(row=row, column=6).number_format = "0.00"

        # Reasonable column widths.
        set_col_widths(
            ws,
            {
                1: 6,   # Row
                2: 14,  # Metric
                3: 16,  # central
                4: 16,  # symbolic
                5: 16,  # diff
                6: 12,  # ratio
            },
        )

        case_index += 1

    # If no sheets were created, add a placeholder sheet to satisfy Excel.
    if not wb.sheetnames:
        ws = wb.create_sheet(title="no-cases")
        ws.cell(row=1, column=1, value="No valid case folders found.")

    # Ensure output directory exists.
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote workbook: {out_path}")


if __name__ == "__main__":
    main()
