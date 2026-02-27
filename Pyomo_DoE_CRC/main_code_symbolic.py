#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import re
import sys
import subprocess
from pathlib import Path
from openpyxl import Workbook

PYTHON = sys.executable

# (sheet_name, script_path, extra_args)
CASES = [
    ("case1_reactor",   "4_state_reactor_sym.py", []),
    ("case2_2paramsin", "two_param_sin_sym.py",   ["--wdir"]),
    ("case3_scalarPDE", "PDE_diffusion_sym.py",   ["--wdir"]),
    ("case4_6param",    "4st_6pmt_sym.py",        ["--wdir"]),
]

N_RUNS = 10
OUT_XLSX = Path("doe_benchmark_symbolic.xlsx")
LOG_DIR = Path("results")


# ---------- last-match helpers ----------
def grab_float(text: str, pat: str, flags: int = 0):
    ms = re.findall(pat, text, flags)
    return float(ms[-1]) if ms else None

def grab_int(text: str, pat: str, flags: int = 0):
    ms = re.findall(pat, text, flags)
    return int(ms[-1]) if ms else None

def grab_str(text: str, pat: str, flags: int = 0):
    ms = re.findall(pat, text, flags)
    return ms[-1].strip() if ms else None

def grab_pair(text: str, pat: str, flags: int = 0):
    ms = re.findall(pat, text, flags)
    if not ms:
        return None, None
    a, b = ms[-1]
    return float(a), float(b)


def parse_stdout_metrics(blob: str) -> dict:
    d = {}

    # IPOPT footer: iteration/eval counts/timing/exit (LAST occurrence)
    d["ipopt_iterations"] = grab_int(blob, r"Number of Iterations\.*:\s*(\d+)")
    d["ipopt_obj_eval"] = grab_int(blob, r"Number of objective function evaluations\s*=\s*(\d+)")
    d["ipopt_grad_eval"] = grab_int(blob, r"Number of objective gradient evaluations\s*=\s*(\d+)")
    d["ipopt_eq_con_eval"] = grab_int(blob, r"Number of equality constraint evaluations\s*=\s*(\d+)")
    d["ipopt_ineq_con_eval"] = grab_int(blob, r"Number of inequality constraint evaluations\s*=\s*(\d+)")
    d["ipopt_eq_jac_eval"] = grab_int(blob, r"Number of equality constraint Jacobian evaluations\s*=\s*(\d+)")
    d["ipopt_ineq_jac_eval"] = grab_int(blob, r"Number of inequality constraint Jacobian evaluations\s*=\s*(\d+)")
    d["ipopt_hess_eval"] = grab_int(blob, r"Number of Lagrangian Hessian evaluations\s*=\s*(\d+)")
    d["ipopt_cpu_no_eval_s"] = grab_float(
        blob,
        r"Total CPU secs in IPOPT \(w/o function evaluations\)\s*=\s*([0-9.eE+-]+)",
    )
    d["ipopt_cpu_nlp_eval_s"] = grab_float(
        blob,
        r"Total CPU secs in NLP function evaluations\s*=\s*([0-9.eE+-]+)",
    )
    d["ipopt_exit"] = grab_str(blob, r"EXIT:\s*([^\r\n]+)")

    # IPOPT end-of-solve diagnostics (scaled/unscaled) — LAST occurrence
    d["ipopt_final_obj_scaled"], d["ipopt_final_obj_unscaled"] = grab_pair(
        blob, r"Objective\.*:\s*([-+0-9.eE]+)\s+([-+0-9.eE]+)"
    )
    d["ipopt_dual_inf_scaled"], d["ipopt_dual_inf_unscaled"] = grab_pair(
        blob, r"Dual infeasibility\.*:\s*([-+0-9.eE]+)\s+([-+0-9.eE]+)"
    )
    d["ipopt_constr_viol_scaled"], d["ipopt_constr_viol_unscaled"] = grab_pair(
        blob, r"Constraint violation\.*:\s*([-+0-9.eE]+)\s+([-+0-9.eE]+)"
    )
    d["ipopt_bound_viol_scaled"], d["ipopt_bound_viol_unscaled"] = grab_pair(
        blob, r"Variable bound violation\.*:\s*([-+0-9.eE]+)\s+([-+0-9.eE]+)"
    )
    d["ipopt_compl_scaled"], d["ipopt_compl_unscaled"] = grab_pair(
        blob, r"Complementarity\.*:\s*([-+0-9.eE]+)\s+([-+0-9.eE]+)"
    )
    d["ipopt_overall_nlp_err_scaled"], d["ipopt_overall_nlp_err_unscaled"] = grab_pair(
        blob, r"Overall NLP error\.*:\s*([-+0-9.eE]+)\s+([-+0-9.eE]+)"
    )

    # DoE timings / objective (from your prints) — LAST occurrence
    d["solve_time_s"] = grab_float(blob, r"Solve time \(s\):\s*([0-9.eE+-]+)")
    d["build_time_s"] = grab_float(blob, r"Build time \(s\):\s*([0-9.eE+-]+)")
    d["init_time_s"]  = grab_float(blob, r"Initialization time \(s\):\s*([0-9.eE+-]+)")
    d["wall_time_s"]  = grab_float(blob, r"Total wall time \(s\):\s*([0-9.eE+-]+)")
    d["objective_value"] = grab_float(blob, r"Objective value at optimal design:\s*([0-9.eE+-]+)")

    # Generic, case-agnostic design capture (store as one cell) — LAST occurrence
    m = re.findall(r"Optimal experiment values:\s*(.*?)(?:FIM at optimal design:|\Z)", blob, re.DOTALL)
    d["design_vector_raw"] = m[-1].strip() if m else None

    return d


def main():
    LOG_DIR.mkdir(exist_ok=True)

    # start clean each time
    if OUT_XLSX.exists():
        OUT_XLSX.unlink()

    wb = Workbook()
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        wb.remove(wb["Sheet"])

    headers = [
        "run_id",
        "returncode",
        "stdout_len",
        "stderr_len",
        "stderr_head",

        "ipopt_exit",
        "ipopt_iterations",
        "ipopt_obj_eval",
        "ipopt_grad_eval",
        "ipopt_eq_con_eval",
        "ipopt_ineq_con_eval",
        "ipopt_eq_jac_eval",
        "ipopt_ineq_jac_eval",
        "ipopt_hess_eval",
        "ipopt_cpu_no_eval_s",
        "ipopt_cpu_nlp_eval_s",

        "ipopt_final_obj_scaled",
        "ipopt_final_obj_unscaled",
        "ipopt_dual_inf_scaled",
        "ipopt_dual_inf_unscaled",
        "ipopt_constr_viol_scaled",
        "ipopt_constr_viol_unscaled",
        "ipopt_bound_viol_scaled",
        "ipopt_bound_viol_unscaled",
        "ipopt_compl_scaled",
        "ipopt_compl_unscaled",
        "ipopt_overall_nlp_err_scaled",
        "ipopt_overall_nlp_err_unscaled",

        "solve_time_s",
        "build_time_s",
        "init_time_s",
        "wall_time_s",
        "objective_value",
        "design_vector_raw",
    ]

    for sheet_name, script, extra_args in CASES:
        script_path = Path(script).resolve()
        if not script_path.exists():
            raise FileNotFoundError(f"Missing case script: {script_path}")

        # IMPORTANT FIX: run each case with its own working directory (like Spyder --wdir)
        script_dir = script_path.parent

        ws = wb.create_sheet(sheet_name)
        ws.append(headers)

        for run_id in range(1, N_RUNS + 1):
            cmd = [PYTHON, str(script_path)] + list(extra_args)
            proc = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                cwd=str(script_dir),  # <<<<<< KEY FIX
            )

            stdout = proc.stdout or ""
            stderr = proc.stderr or ""
            blob = stdout + "\n" + stderr

            # save full log for this run
            (LOG_DIR / f"{sheet_name}_run{run_id:03d}.log").write_text(
                blob, encoding="utf-8", errors="replace"
            )

            metrics = parse_stdout_metrics(blob)
            stderr_head = (stderr[:150].replace("\n", " ")) if stderr else ""

            row = [
                run_id,
                proc.returncode,
                len(stdout),
                len(stderr),
                stderr_head,

                metrics.get("ipopt_exit"),
                metrics.get("ipopt_iterations"),
                metrics.get("ipopt_obj_eval"),
                metrics.get("ipopt_grad_eval"),
                metrics.get("ipopt_eq_con_eval"),
                metrics.get("ipopt_ineq_con_eval"),
                metrics.get("ipopt_eq_jac_eval"),
                metrics.get("ipopt_ineq_jac_eval"),
                metrics.get("ipopt_hess_eval"),
                metrics.get("ipopt_cpu_no_eval_s"),
                metrics.get("ipopt_cpu_nlp_eval_s"),

                metrics.get("ipopt_final_obj_scaled"),
                metrics.get("ipopt_final_obj_unscaled"),
                metrics.get("ipopt_dual_inf_scaled"),
                metrics.get("ipopt_dual_inf_unscaled"),
                metrics.get("ipopt_constr_viol_scaled"),
                metrics.get("ipopt_constr_viol_unscaled"),
                metrics.get("ipopt_bound_viol_scaled"),
                metrics.get("ipopt_bound_viol_unscaled"),
                metrics.get("ipopt_compl_scaled"),
                metrics.get("ipopt_compl_unscaled"),
                metrics.get("ipopt_overall_nlp_err_scaled"),
                metrics.get("ipopt_overall_nlp_err_unscaled"),

                metrics.get("solve_time_s"),
                metrics.get("build_time_s"),
                metrics.get("init_time_s"),
                metrics.get("wall_time_s"),
                metrics.get("objective_value"),
                metrics.get("design_vector_raw"),
            ]

            ws.append(row)
            wb.save(OUT_XLSX)

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
